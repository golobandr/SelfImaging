import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def setSheetNames(cols, col_names, col_widths, wb, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb.active(sheet_name=sheet_name)
    else:
        ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 35
    for i in range(1, cols + 1):
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.cell(row=1, column=i).value = col_names[i - 1]
        ws.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = col_widths[col[0].col_idx - 1]
    return ws


def fromDatLine(result):
    def printoutCellData(is_ok):
        if is_ok:
            return 'OK'
        else:
            return 'NOK'

    wb = load_workbook(result.io.outputfile)
    col_names = ['Status', 'Start time', 'Finish time', 'Message']
    col_widths = [12, 20, 20, 80]
    ws = setSheetNames(len(col_names), col_names, col_widths, wb, 'Result')
    for idx in range(len(result.data)):
        ws.append([printoutCellData(result.data[idx].is_ok), result.data[idx].start,
                   result.data[idx].end, result.data[idx].message])
    if result.is_ok and 'distance' in result.dependencies:
        col_names = ['', 'z, mm']
        for idx in range(len(result.data)):
            col_names.append(result.data[idx].psd.distance)
        if 'Distance' in wb.sheetnames:
            ws = wb.active
        else:
            ws = wb.create_sheet('Distance')
        ws.append(col_names)
        ws.append(['x, mm'])
        for i in range(len(result.data[0].psd.image.x.coordinate)):
            data_to_save = [result.data[0].psd.image.x.coordinate[i], '']
            for idx in range(len(result.data)):
                data_to_save.append(result.data[idx].psd.image.x.intensity[i])
            ws.append(data_to_save)

    for idx in range(len(result.data)):
        if result.data[idx].is_ok and result.data[idx].add.save:
            if result.data[idx].add.debug:
                col_names = ['x frequencies', 'x spectrum, urb.un.', 'y frequencies', 'y spectrum, urb.un.',
                             'x, mm', 'x intensities, urb.un.', 'y, mm', 'y intensities, urb.un.', ]
                col_widths = [15, 21, 15, 21, 12, 25, 12, 25]
                rows = max([len(result.data[idx].grating.coefficients.x.n),
                            len(result.data[idx].grating.coefficients.y.n),
                            len(result.data[idx].psd.image.x.coordinate),
                            len(result.data[idx].psd.image.y.coordinate)])
            else:
                if '1D' in result.data[idx].grating.slit:
                    col_names = ['x', 'x intensities, urb.un.']
                    col_widths = [12, 25]
                    rows = max(
                        [len(result.data[idx].psd.image.x.coordinate), len(result.data[idx].psd.image.y.coordinate)])
                else:
                    col_names = ['x', 'x intensities, urb.un.', 'y', 'y intensities, urb.un.', ]
                    col_widths = [12, 25, 12, 25]
                    rows = max(
                        [len(result.data[idx].psd.image.x.coordinate), len(result.data[idx].psd.image.y.coordinate)])
            ws = setSheetNames(len(col_names), col_names, col_widths, wb, f'Line #{idx}')
            for i in range(rows):
                data_to_save = []
                if result.data[idx].add.debug:
                    if len(result.data[idx].grating.coefficients.x.n) > i:
                        data_to_save.append(result.data[idx].grating.coefficients.x.n[i])
                        data_to_save.append(result.data[idx].grating.coefficients.x.sn[i])
                    else:
                        data_to_save.append('')
                        data_to_save.append('')
                    if len(result.data[idx].grating.coefficients.y.n) > i:
                        data_to_save.append(result.data[idx].grating.coefficients.y.n[i])
                        data_to_save.append(result.data[idx].grating.coefficients.y.sn[i])
                    else:
                        data_to_save.append('')
                        data_to_save.append('')
                if len(result.data[idx].psd.image.x.coordinate) > i:
                    data_to_save.append(result.data[idx].psd.image.x.coordinate[i])
                    data_to_save.append(result.data[idx].psd.image.x.intensity[i])
                else:
                    data_to_save.append('')
                    data_to_save.append('')
                if len(result.data[idx].psd.image.y.coordinate) > i or not '1D' in result.data[idx].grating.slit:
                    data_to_save.append(result.data[idx].psd.image.y.coordinate[i])
                    data_to_save.append(result.data[idx].psd.image.y.intensity[i])
                else:
                    data_to_save.append('')
                    data_to_save.append('')
                ws.append(data_to_save)
    wb.save(result.io.outputfile)


def forDependencies(result):
    to_save = False
    wb = load_workbook(result.io.outputfile)
    if 'error' in result.dependencies.keys():
        error = result.dependencies['error']
        col_names = ['Step division', 'Error']
        col_widths = [15, 15]
        ws = setSheetNames(len(col_names), col_names, col_widths, wb, 'CalculationError')
        for idx in range(len(error['n'])):
            ws.append([error['n'][idx], error['data'][idx]])
        to_save = True
    if to_save:
        wb.save(result.io.outputfile)
